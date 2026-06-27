"""Daily editable projection workbook (.xlsx).

Project 54.7 ships a model win probability for every game market and prop. The
*price* you can get changes all day and differs by book, so this workbook makes
the **odds the only thing you edit**: the model number is locked, you type the
price your sportsbook is showing into the yellow cells, and the edge, expected
value, and quarter-Kelly stake recompute in-cell — in Excel *or* Google Sheets,
with no internet connection and nothing to go stale.

That's the whole idea, and it's the opposite of every paid tool in the space:
they de-vig a sharp market for you and lock the math in a live web app you can't
touch (and whose lines are the #1 thing users complain go stale). Here the math
is in front of you and the inputs are yours.

Design rules baked in (from competitive research; see docs/WORKBOOK.md):
  * Editable cells are unmistakable by *sight* (yellow fill + border + header),
    not just by sheet protection — because opening a protected .xlsx in Google
    Sheets silently drops the protection, so locking can never be the only cue.
  * No macros, no Power Query, no live web imports (IMPORTXML can't read a
    JS-rendered sportsbook anyway) — everything is plain values + formulas that
    survive the Excel <-> Sheets round trip.
  * The math (American->implied, EV, Kelly) mirrors project547.odds exactly, so
    the sheet and the engine never disagree.

Pure functions over a loaded ``latest.json`` dict, so the whole thing is
unit-tested without a spreadsheet app. Entry point: :func:`build_workbook`.
"""

from __future__ import annotations

import io
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from . import config, odds

# ---------------------------------------------------------------------------
# Palette (echoes the dashboard's cream/graphite brand, but printable)
# ---------------------------------------------------------------------------
INK = "1F2328"
MUTED = "5A6066"
EDIT = "FFE9A8"        # the yellow "type here" fill
EDIT_BORDER = "C8941A"
HEAD_BG = "1F2328"
HEAD_FG = "FBF5E6"
BAND = "F4EAD0"        # zebra band
GOOD = "2F7A4A"
NEG = "B03636"
LINE = "D8CCAB"

# Settings cells every formula points at (one edit reprices the whole book).
SET_SHEET = "Settings"
CELL_BANKROLL = f"'{SET_SHEET}'!$B$2"
CELL_KELLY_FRAC = f"'{SET_SHEET}'!$B$3"
CELL_MIN_EDGE = f"'{SET_SHEET}'!$B$4"
CELL_MAX_STAKE = f"'{SET_SHEET}'!$B$5"

DISCLAIMER = ("Personal research. Not financial advice. No bankroll promises. "
              "If gambling stops being fun, stop — call 1-800-GAMBLER.")


# ---------------------------------------------------------------------------
# Small value helpers
# ---------------------------------------------------------------------------
def _num(x):
    """Return a finite float or None (collapses None and NaN/inf)."""
    if x is None:
        return None
    try:
        f = float(x)
    except (TypeError, ValueError):
        return None
    return f if math.isfinite(f) else None


def _fair_american(prob: float | None):
    """The model's own fair price (no vig) for a probability, as American odds."""
    p = _num(prob)
    if p is None or not (0.0 < p < 1.0):
        return None
    return odds.decimal_to_american(1.0 / p)


# ---------------------------------------------------------------------------
# In-cell formula builders (mirror project547.odds; unit-tested directly)
#
# ``o`` is the editable American-odds cell, ``p`` the locked model-prob cell.
# Every formula no-ops to "" while the odds cell is blank so an un-filled row
# never shows a misleading 0% edge.
# ---------------------------------------------------------------------------
def _profit(o: str) -> str:
    """Decimal profit per 1 unit staked at American odds in ``o`` (= dec - 1)."""
    return f"IF({o}<0,-100/{o},{o}/100)"


def implied_formula(o: str) -> str:
    return f'=IF({o}="","",IF({o}<0,-{o}/(-{o}+100),100/({o}+100)))'


def ev_formula(p: str, o: str) -> str:
    return f'=IF({o}="","",{p}*{_profit(o)}-(1-{p}))'


def kelly_pct_formula(p: str, o: str) -> str:
    """Fractional-Kelly stake as a % of bankroll (0 when there's no edge)."""
    b = _profit(o)
    full = f"({p}*{b}-(1-{p}))/{b}"
    return f'=IF({o}="","",MAX(0,{full})*{CELL_KELLY_FRAC})'


def stake_formula(kelly_cell: str, o: str) -> str:
    """Dollars to risk: fractional Kelly, capped at the max-stake setting."""
    return (f'=IF({o}="","",ROUND(MIN({kelly_cell},{CELL_MAX_STAKE})'
            f"*{CELL_BANKROLL},2))")


def verdict_formula(ev_cell: str) -> str:
    """Plain-language read of the EV cell, anti-guru and research-aware.

    The model's *validated* edge lives in the moderate 2-6% EV band (good CLV);
    very high EV on these efficient markets usually means a stale/oddball price,
    not free money, so we say "verify" rather than "smash".
    """
    return (
        f'=IF({ev_cell}="","",'
        f'IF({ev_cell}<0,"Pass",'
        f'IF({ev_cell}<{CELL_MIN_EDGE},"Thin",'
        f'IF({ev_cell}<0.06,"✅ Sharp band",'
        f'IF({ev_cell}<0.08,"Strong",'
        f'"⚠ Verify (stale?)")))))'
    )


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
EDIT_SIDE = Side(style="medium", color=EDIT_BORDER)
EDIT_BOX = Border(left=EDIT_SIDE, right=EDIT_SIDE, top=EDIT_SIDE, bottom=EDIT_SIDE)
EDIT_FILL = PatternFill("solid", fgColor=EDIT)
HEAD_FILL = PatternFill("solid", fgColor=HEAD_BG)
BAND_FILL = PatternFill("solid", fgColor=BAND)
HEAD_FONT = Font(name="DM Sans", bold=True, color=HEAD_FG, size=10)
CELL_FONT = Font(name="DM Sans", color=INK, size=10)
MUTED_FONT = Font(name="DM Sans", color=MUTED, size=10)
UNLOCKED = Protection(locked=False)


def _title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name="Oswald", bold=True, size=18, color=INK)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = MUTED_FONT


def _header_row(ws, row, headers, edit_cols=()):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# ---------------------------------------------------------------------------
# Slate -> uniform market rows
# ---------------------------------------------------------------------------
def game_rows(sport: str, date: str, g: dict) -> list[dict]:
    """Flatten one game into per-side rows: (market, selection, model prob,
    market odds). Sport-agnostic — handles MLB run-line and the generic-engine
    spread fields, and emits a row whenever the model has a probability, even
    if no market price is posted yet (you can still type your book's number)."""
    home, away = g.get("home_team", "Home"), g.get("away_team", "Away")
    matchup = f"{away} @ {home}"
    start = str(g.get("game_time", ""))[11:16]
    out: list[dict] = []

    def add(market, selection, prob, mkt_odds, stored_ev=None, line=None):
        o = _num(mkt_odds)
        # Prefer the probability the engine actually bet on: where it stored an
        # EV against a real price, recover p = (EV+1)/(b+1) so the sheet's math
        # equals the engine's (it shrinks toward market before pricing). With no
        # market EV, fall back to the raw model probability.
        ev, p = _num(stored_ev), _num(prob)
        priced = False
        if ev is not None and o is not None:
            b = odds.american_to_decimal(o) - 1
            if b > 0:
                p = (ev + 1.0) / (b + 1.0)
                priced = True  # the engine actually priced & bets this market
        if p is None or not (0.0 < p < 1.0):
            return
        out.append({"sport": sport, "date": date, "matchup": matchup,
                    "start": start, "market": market, "selection": selection,
                    "prob": p, "odds": o, "line": _num(line),
                    "engine_priced": priced})

    # Moneyline
    add("Moneyline", away, g.get("away_win_prob"), g.get("away_ml"), g.get("away_ml_ev"))
    add("Moneyline", home, g.get("home_win_prob"), g.get("home_ml"), g.get("home_ml_ev"))

    # Total (over/under)
    tl = _num(g.get("total_line"))
    mover = _num(g.get("model_over_prob"))
    if mover is not None:
        ln = f"{tl:g}" if tl is not None else ""
        add("Total", f"Over {ln}".strip(), mover, g.get("over_odds"),
            g.get("over_ev"), tl)
        add("Total", f"Under {ln}".strip(), 1 - mover,
            g.get("under_odds", g.get("over_odds")), g.get("under_ev"), tl)

    # Spread / run line — MLB uses rl_*, generic engine uses spread_*
    sp_line = _num(g.get("rl_home_line", g.get("spread_home_line")))
    home_cover = _num(g.get("model_home_rl", g.get("model_home_cover")))
    h_odds = g.get("rl_home_odds", g.get("spread_home_odds"))
    a_odds = g.get("rl_away_odds", g.get("spread_away_odds"))
    h_ev = g.get("rl_home_ev", g.get("spread_home_ev"))
    a_ev = g.get("rl_away_ev", g.get("spread_away_ev"))
    if home_cover is not None and sp_line is not None:
        add("Spread", f"{home} {sp_line:+g}", home_cover, h_odds, h_ev, sp_line)
        add("Spread", f"{away} {-sp_line:+g}", 1 - home_cover, a_odds, a_ev, -sp_line)
    return out


def prop_rows(sport: str, date: str, props: list[dict]) -> list[dict]:
    """Priced prop rows (Over side) — only those with a posted line, a price,
    and a model over-probability, since those are the ones with a live edge."""
    out = []
    for p in props:
        line, o, mp = _num(p.get("line")), _num(p.get("odds")), _num(p.get("model_over_prob"))
        if line is None or mp is None or not (0.0 < mp < 1.0):
            continue
        mkt = str(p.get("market", "")).replace("_", " ").title()
        out.append({"sport": sport, "date": date, "player": p.get("player", ""),
                    "team": p.get("team", ""), "opp": p.get("opponent", ""),
                    "market": mkt, "line": line, "selection": f"Over {line:g}",
                    "prob": mp, "odds": o})
    return out


def _why(r: dict) -> str:
    """A one-line, anti-guru model read for the Notes column."""
    pct = f"{r['prob'] * 100:.1f}%"
    fair = _fair_american(r["prob"])
    fair_s = f"{fair:+d}" if fair is not None else "n/a"
    return f"Model {pct} (fair {fair_s}). Your call."


# ---------------------------------------------------------------------------
# Sheet renderers
# ---------------------------------------------------------------------------
GAME_HEADERS = ["Date", "Matchup", "Start", "Market", "Selection",
                "Model Win%", "Model Fair", "✏ Your Odds", "Implied%",
                "Your Edge (EV)", "¼-Kelly %", "Stake $", "Verdict", "Model read"]
# column letters for the formula-bearing columns
_C_PROB, _C_ODDS, _C_IMP, _C_EV, _C_KPCT, _C_STAKE, _C_VERD = "F", "H", "I", "J", "K", "L", "M"


def _write_market_table(ws, rows, start_row, show_player=False):
    """Shared renderer for game/top-plays/prop tables with live formulas."""
    headers = GAME_HEADERS
    if show_player:
        headers = (["Date", "Player", "Team", "Opp", "Market", "Selection",
                    "Model Over%", "Model Fair", "✏ Your Odds", "Implied%",
                    "Your Edge (EV)", "¼-Kelly %", "Stake $", "Verdict"])
    _header_row(ws, start_row, headers)
    r = start_row + 1
    for i, row in enumerate(rows):
        if show_player:
            prob_col, odds_col = "G", "I"
            imp_col, ev_col, k_col, stake_col, verd_col = "J", "K", "L", "M", "N"
            vals = [row["date"], row["player"], row["team"], row["opp"],
                    row["market"], row["selection"]]
            base = len(vals)
        else:
            prob_col, odds_col = _C_PROB, _C_ODDS
            imp_col, ev_col, k_col, stake_col, verd_col = _C_IMP, _C_EV, _C_KPCT, _C_STAKE, _C_VERD
            vals = [row["date"], row["matchup"], row.get("start", ""),
                    row["market"], row["selection"]]
            base = len(vals)
        # static cells
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = BORDER
            if i % 2:
                cell.fill = BAND_FILL
        P = f"${prob_col}${r}"
        O = f"${odds_col}${r}"
        K = f"${k_col}${r}"
        EV = f"${ev_col}${r}"
        # model prob (locked) + fair price (locked)
        pc = ws[f"{prob_col}{r}"]; pc.value = row["prob"]; pc.number_format = "0.0%"
        fc_col = "H" if show_player else "G"
        fair = _fair_american(row["prob"])
        fc = ws[f"{fc_col}{r}"]; fc.value = fair if fair is not None else "—"
        fc.number_format = "+0;-0"
        # editable odds cell
        oc = ws[f"{odds_col}{r}"]
        oc.value = row.get("odds")  # default to the market price if we have it
        oc.fill = EDIT_FILL
        oc.border = EDIT_BOX
        oc.protection = UNLOCKED
        oc.number_format = "+0;-0"
        oc.alignment = Alignment(horizontal="center")
        # formulas
        ws[f"{imp_col}{r}"] = implied_formula(O); ws[f"{imp_col}{r}"].number_format = "0.0%"
        ws[f"{ev_col}{r}"] = ev_formula(P, O); ws[f"{ev_col}{r}"].number_format = "+0.0%;-0.0%"
        ws[f"{k_col}{r}"] = kelly_pct_formula(P, O); ws[f"{k_col}{r}"].number_format = "0.00%"
        ws[f"{stake_col}{r}"] = stake_formula(K, O); ws[f"{stake_col}{r}"].number_format = "$#,##0.00"
        ws[f"{verd_col}{r}"] = verdict_formula(EV)
        if not show_player:
            wc = ws.cell(row=r, column=14, value=_why(row)); wc.font = MUTED_FONT
        # borders on every cell; default font on the computed (non-static) ones
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = EDIT_BOX if c == ws[f"{odds_col}{r}"].column else BORDER
            if c > base and c != ws[f"{odds_col}{r}"].column and c != 14:
                cell.font = CELL_FONT
        r += 1
    # green->red scale on the EV column so edges read at a glance
    if r > start_row + 1:
        rng = f"{ev_col}{start_row + 1}:{ev_col}{r - 1}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=-0.05, start_color=NEG,
            mid_type="num", mid_value=0, mid_color="FBF5E6",
            end_type="num", end_value=0.08, end_color=GOOD))
    ws.protection.sheet = True
    ws.protection.password = ""        # cosmetic lock only; inputs stay yellow
    return r


def _autosize(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


_GAME_WIDTHS = {"A": 11, "B": 30, "C": 7, "D": 11, "E": 22, "F": 10, "G": 9,
                "H": 11, "I": 9, "J": 12, "K": 10, "L": 10, "M": 16, "N": 30}
_PROP_WIDTHS = {"A": 11, "B": 20, "C": 8, "D": 8, "E": 18, "F": 14, "G": 11,
                "H": 9, "I": 11, "J": 9, "K": 12, "L": 10, "M": 10, "N": 16}


def build_readme(wb, data):
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    _title(ws, "🎯 Project 54.7 — Daily Wager Workbook",
           "52.4% pays the house. 54.7% pays you.")
    gen = str(data.get("generated_at", ""))[:16].replace("T", " ")
    lines = [
        ("", ""),
        ("Built", gen),
        ("", ""),
        ("HOW IT WORKS", ""),
        ("1.", "The model's win probability for every market is locked. "
               "It does not change when odds move."),
        ("2.", "Type the price YOUR sportsbook is showing into the yellow "
               "“✏ Your Odds” cells (American odds, e.g. -110 or +135)."),
        ("3.", "Implied %, Your Edge (EV), the ¼-Kelly stake and the Verdict "
               "recompute instantly — in Excel or Google Sheets, offline."),
        ("4.", "Set your Bankroll and Kelly fraction once on the Settings tab; "
               "every stake on every tab follows it."),
        ("", ""),
        ("READING THE VERDICT", ""),
        ("Pass", "Model says there's no edge at your price — skip it."),
        ("Thin", "Positive but below your min-edge threshold."),
        ("✅ Sharp band", "2–6% EV — where the model's edge has actually held up "
                          "(good closing-line value). The sweet spot."),
        ("Strong", "6–8% EV — bigger, rarer."),
        ("⚠ Verify", "8%+ EV usually means a stale or oddball price, not free "
                     "money. Double-check the line before trusting it."),
        ("", ""),
        ("WHY EDIT YOUR OWN ODDS?", ""),
        ("", "Every other tool de-vigs a sharp book for you and locks the math "
             "in an app whose lines go stale. Here the number is yours, so "
             "nothing can be out of date — and the math is in front of you."),
        ("", ""),
        ("THE DEAL", DISCLAIMER),
        ("", "We publish our losses too — see the Track Record tab. We tell you "
             "when to pass. The trigger is always yours."),
    ]
    r = 4
    for k, v in lines:
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        a.font = Font(name="DM Sans", bold=bool(k and k.isupper()), color=INK, size=10)
        b.font = CELL_FONT
        b.alignment = Alignment(wrap_text=True, vertical="top")
        if k.isupper() and v == "":
            a.font = Font(name="Oswald", bold=True, color=GOOD, size=12)
        r += 1
    _autosize(ws, {"A": 16, "B": 86})
    ws.protection.sheet = True
    return ws


def build_settings(wb):
    ws = wb.create_sheet(SET_SHEET)
    ws.sheet_view.showGridLines = False
    _title(ws, "⚙ Settings", "Edit the yellow cells — every tab reprices.")
    rows = [
        ("Bankroll ($)", 1000, "$#,##0"),
        ("Kelly fraction", 0.25, "0.00"),
        ("Min edge to flag (EV)", config.MIN_EDGE, "0.0%"),
        ("Max stake (% of bankroll)", 0.05, "0.0%"),
    ]
    r = 2
    for label, val, fmt in rows:
        lc = ws.cell(row=r, column=1, value=label); lc.font = CELL_FONT
        vc = ws.cell(row=r, column=2, value=val)
        vc.fill = EDIT_FILL; vc.border = EDIT_BOX; vc.protection = UNLOCKED
        vc.number_format = fmt; vc.alignment = Alignment(horizontal="center")
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Default ¼-Kelly is deliberate: full Kelly is mathematically "
                  "optimal but brutal on variance, and edges are always "
                  "overestimated. Quarter-Kelly keeps ~most of the growth at a "
                  "fraction of the risk.").font = MUTED_FONT
    ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=2)
    _autosize(ws, {"A": 30, "B": 16})
    ws.protection.sheet = True
    return ws


def build_top_plays(wb, all_rows, min_edge=None, cap=25):
    """Curated highest-edge plays across every sport, ranked by the model's EV
    at the posted price. Light enough to glance at on a phone."""
    min_edge = config.MIN_EDGE if min_edge is None else min_edge
    scored = []
    for row in all_rows:
        o = row.get("odds")
        # only markets the engine actually priced & bets — never raw-model fat
        # edges on markets it deliberately skips (those are stale, not value)
        if o is None or not row.get("engine_priced"):
            continue
        ev = odds.expected_value(row["prob"], o)
        if ev >= min_edge:
            scored.append((ev, row))
    scored.sort(key=lambda x: x[0], reverse=True)
    rows = [r for _, r in scored[:cap]]

    ws = wb.create_sheet("Top Plays")
    _title(ws, "⭐ Top Plays — biggest model edges today",
           "Ranked at the posted price. Swap in your book's number and watch "
           "the edge update. Nothing here is a lock.")
    if not rows:
        ws.cell(row=4, column=1,
                value="No qualifying edges at the posted prices right now. "
                      "Open a sport tab and type in your book's odds — an edge "
                      "the market hasn't is exactly what you're hunting.").font = MUTED_FONT
        ws.protection.sheet = True
        _autosize(ws, _GAME_WIDTHS)
        return ws
    _write_market_table(ws, rows, start_row=4)
    _autosize(ws, _GAME_WIDTHS)
    return ws


def build_sport_sheets(wb, sport, game_rows_, prop_rows_):
    if game_rows_:
        ws = wb.create_sheet(f"{sport} Games")
        _title(ws, f"{sport} — game markets",
               "Moneyline · Total · Spread. Yellow = your price.")
        _write_market_table(ws, game_rows_, start_row=4)
        _autosize(ws, _GAME_WIDTHS)
    if prop_rows_:
        ws = wb.create_sheet(f"{sport} Props")
        _title(ws, f"{sport} — player props",
               "Over side priced off the model. Yellow = your price.")
        _write_market_table(ws, prop_rows_, start_row=4, show_player=True)
        _autosize(ws, _PROP_WIDTHS)


def build_track_record(wb, data):
    perf = data.get("performance") or {}
    ws = wb.create_sheet("Track Record")
    ws.sheet_view.showGridLines = False
    _title(ws, "📒 Track Record — receipts, losses included",
           "Graded in the open. A process you can audit beats a hot streak you "
           "can't.")
    headers = ["Scope", "Graded games", "Model Brier", "Bets", "Win %",
               "Units", "ROI %", "Avg CLV %", "CLV beat %"]
    _header_row(ws, 4, headers)
    keymap = [("graded_games", "0"), ("model_brier", "0.000"), ("bets", "0"),
              ("bet_win_rate", "0.0%"), ("units", "+0.00;-0.00"),
              ("roi_pct", "+0.0;-0.0"), ("avg_clv_pct", "+0.0;-0.0"),
              ("clv_beat_rate", "0.0%")]
    r = 5

    def emit(label, blob):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = CELL_FONT
        for c, (k, fmt) in enumerate(keymap, start=2):
            v = blob.get(k)
            # win/beat rates are stored 0-1; clv/roi are already %
            cell = ws.cell(row=r, column=c, value=_num(v))
            cell.number_format = fmt
            cell.font = CELL_FONT
            cell.border = BORDER
        ws.cell(row=r, column=1).border = BORDER
        r += 1

    if perf.get("overall"):
        emit("Overall", perf["overall"])
    for sp, blob in (perf.get("by_sport") or {}).items():
        emit(sp, blob)
    if r == 5:
        ws.cell(row=5, column=1,
                value="No graded results yet — the record starts filling in as "
                      "today's projections settle.").font = MUTED_FONT
    note = ("Brier: lower is better (0.25 = a coin flip). CLV beat %: how often "
            "we beat the closing line — the earliest honest sign an edge is "
            "real, before win/loss catches up.")
    ws.cell(row=r + 1, column=1, value=note).font = MUTED_FONT
    ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=9)
    _autosize(ws, {"A": 14, "B": 13, "C": 11, "D": 8, "E": 9, "F": 9,
                   "G": 9, "H": 10, "I": 11})
    ws.protection.sheet = True
    return ws


# ---------------------------------------------------------------------------
# Research Hub — premium matchup pages (image card + all-windows table)
# ---------------------------------------------------------------------------
RANK_GOOD = PatternFill("solid", fgColor="D6EAD9")
RANK_BAD = PatternFill("solid", fgColor="F1D6D6")
RANK_MID = PatternFill("solid", fgColor="F3E7C8")
HUB_WINDOWS = ["season", "l30", "l20", "l15", "l10", "l5"]
HUB_WIN_LABELS = {"season": "Season", "l30": "L30", "l20": "L20", "l15": "L15",
                  "l10": "L10", "l5": "L5"}


def _safe_sheet_name(name: str, used: set) -> str:
    bad = set(r'[]:*?/\\')
    s = "".join(c for c in name if c not in bad)[:28].strip() or "Game"
    base, i = s, 2
    while s in used:
        s = f"{base[:25]} {i}"
        i += 1
    used.add(s)
    return s


def _rank_fill(rank, n):
    if rank is None or not n:
        return None
    if rank <= n / 3:
        return RANK_GOOD
    if rank > 2 * n / 3:
        return RANK_BAD
    return RANK_MID


def _native_matchup_table(ws, start_row, title, rows, n):
    """All-windows offense-vs-defense table (Season..L5 + rank), conditionally
    shaded by rank so any recency window is readable in the static file."""
    tcell = ws.cell(row=start_row, column=1, value=title)
    tcell.font = Font(name="Oswald", bold=True, size=11, color=INK)
    hdr = start_row + 1
    labels = [HUB_WIN_LABELS[w] for w in HUB_WINDOWS]
    headers = (["Stat"] + [f"A {l}" for l in labels] + ["A Rk", "Adv", "H Rk"]
               + [f"H {l}" for l in reversed(labels)] + ["Opp stat"])
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    r = hdr + 1
    for row in rows:
        c = 1

        def put(val, fill=None, bold=False):
            nonlocal c
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="DM Sans", size=9, bold=bold, color=INK)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
            c += 1

        put(row["stat"], bold=True)
        for w in HUB_WINDOWS:
            put(_round(row.get(f"off_{w}")))
        put(_ord_txt(row.get("off_rank")), _rank_fill(row.get("off_rank"), n), bold=True)
        put("★" * int(row.get("adv") or 0))
        put(_ord_txt(row.get("def_rank")), _rank_fill(row.get("def_rank"), n), bold=True)
        for w in reversed(HUB_WINDOWS):
            put(_round(row.get(f"def_{w}")))
        put(f"Opp {row['stat']}", bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        r += 1
    return r + 1


def _round(v):
    f = _num(v)
    return round(f, 3) if f is not None else "—"


def _ord_txt(n):
    if n is None:
        return "—"
    n = int(n)
    suf = "th" if 10 <= n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def build_research_hub(wb, data, render_cards=False, window="l5", max_games=10):
    """Add the Research Hub: an index + one premium matchup page per game
    (embedded card image when a browser is available, plus an all-windows
    table). Fully best-effort — a missing dependency or a single bad game is
    skipped, never fatal. Returns the number of game pages added."""
    try:
        from . import teamstats
    except Exception:  # noqa: BLE001 — pandas/history not available
        return 0
    try:
        from . import cardimage
    except Exception:  # noqa: BLE001
        cardimage = None

    slates = data.get("slates") or {}
    primary = data.get("primary_date") or (data.get("dates") or [None])[0]
    blob = slates.get(primary) or {}
    games = []
    for sport, sb in blob.items():
        for g in (sb.get("games") or []):
            if g.get("home_team") and g.get("away_team"):
                games.append((sport, g))
    if not games:
        return 0
    games = games[:max_games]

    index = wb.create_sheet("Research Hub")
    index.sheet_view.showGridLines = False
    _title(index, "🔬 Research Hub", f"Premium matchup breakdowns · {primary} · "
           f"{teamstats.WINDOW_LABELS.get(window, window)} window. Each game has "
           "its own tab.")
    idx_r = 4
    used = {ws.title for ws in wb.worksheets}
    added = 0
    for sport, g in games:
        home, away = g["home_team"], g["away_team"]
        try:
            m = teamstats.matchup(sport, home, away, primary, window=window)
        except Exception:  # noqa: BLE001
            m = {}
        if not m:
            continue
        title = f"{away.split()[-1]} @ {home.split()[-1]}"
        name = _safe_sheet_name(f"{sport} {title}", used)
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        _title(ws, f"{away} @ {home}",
               f"{sport} · {primary} · {teamstats.WINDOW_LABELS.get(window, window)} window")
        row = 4
        if render_cards and cardimage is not None:
            try:
                png = cardimage.render_matchup_png(sport, g, m, window=window)
            except Exception:  # noqa: BLE001
                png = None
            if png:
                from openpyxl.drawing.image import Image as XLImage
                # openpyxl reads the image lazily at save(), so a BytesIO/closed
                # PIL stream fails. Write each PNG to a temp file (kept alive on
                # the workbook until save) and pass the path.
                tmpdir = getattr(wb, "_osp_card_dir", None)
                if tmpdir is None:
                    import tempfile
                    tmpdir = tempfile.TemporaryDirectory(prefix="osp_cards_")
                    wb._osp_card_dir = tmpdir  # survives until wb is GC'd
                fp = Path(tmpdir.name) / f"{name}.png".replace(" ", "_")
                fp.write_bytes(png)
                img = XLImage(str(fp))
                scale = 0.5 * 0.92  # back out the 2x device scale, fit the page
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
                ws.add_image(img, f"A{row}")
                row += max(20, int(img.height / 18)) + 2
        n = m.get("n_teams", 30)
        row = _native_matchup_table(ws, row, f"{away} offense vs {home} defense",
                                    m.get("away_off_vs_home_def") or [], n)
        row = _native_matchup_table(ws, row, f"{home} offense vs {away} defense",
                                    m.get("home_off_vs_away_def") or [], n)
        ws.protection.sheet = True
        # index link
        ic = index.cell(row=idx_r, column=1, value=f"{sport}: {away} @ {home}")
        ic.hyperlink = f"#'{name}'!A1"
        ic.font = Font(name="DM Sans", size=10, color="1F6FEB", underline="single")
        idx_r += 1
        added += 1
    index.cell(row=idx_r + 1, column=1,
               value="Tip: the website's research view lets you flip the L5/L10/"
                     "L15/L20/L30/Season window live; here every window is in the "
                     "table so you can read any of them.").font = MUTED_FONT
    _autosize(index, {"A": 60})
    return added


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def collect_rows(data):
    """Return (per-sport game rows, per-sport prop rows, flat all-game rows)."""
    slates = data.get("slates") or {}
    dates = data.get("dates") or sorted(slates.keys())
    games_by_sport: dict[str, list] = {}
    props_by_sport: dict[str, list] = {}
    all_game_rows: list = []
    for date in dates:
        for sport, blob in (slates.get(date) or {}).items():
            for g in blob.get("games", []) or []:
                rows = game_rows(sport, date, g)
                games_by_sport.setdefault(sport, []).extend(rows)
                all_game_rows.extend(rows)
            pr = prop_rows(sport, date, blob.get("props", []) or [])
            if pr:
                props_by_sport.setdefault(sport, []).extend(pr)
    return games_by_sport, props_by_sport, all_game_rows


def build_workbook(data: dict, render_cards: bool = False,
                   include_hub: bool = False, window: str = "l5") -> Workbook:
    """Build the full editable workbook from a loaded ``latest.json`` dict.

    ``include_hub`` adds the Research Hub (matchup pages; needs box-score logs
    via teamstats, so it's off for the hermetic on-demand/test path and on for
    disk builds). ``render_cards`` additionally embeds the premium card images
    (needs a browser; slow, so the hourly job sets it). ``window`` picks the
    recency window for the card images / rank coloring."""
    wb = Workbook()
    games_by_sport, props_by_sport, all_game_rows = collect_rows(data)

    build_readme(wb, data)
    build_settings(wb)
    build_top_plays(wb, all_game_rows)
    # stable, in-season-first sport order
    order = ["MLB", "WNBA", "NBA", "NFL", "NCAAF", "NHL"]
    sports = sorted(set(games_by_sport) | set(props_by_sport),
                    key=lambda s: (order.index(s) if s in order else 99, s))
    for sport in sports:
        build_sport_sheets(wb, sport, games_by_sport.get(sport, []),
                           props_by_sport.get(sport, []))
    build_track_record(wb, data)
    if include_hub or render_cards:
        try:
            build_research_hub(wb, data, render_cards=render_cards, window=window)
        except Exception:  # noqa: BLE001 — the hub is a bonus, never block the build
            pass
    return wb


def build_workbook_bytes(data: dict, render_cards: bool = False,
                         include_hub: bool = False) -> bytes:
    """In-memory .xlsx bytes — what the dashboard download button serves."""
    buf = io.BytesIO()
    build_workbook(data, render_cards=render_cards, include_hub=include_hub).save(buf)
    return buf.getvalue()


def default_paths():
    out = config.OUTPUT_DIR / "workbook"
    return out / "latest.xlsx", out


def build_to_disk(data: dict, primary_date: str | None = None,
                  render_cards: bool = True) -> Path:
    """Write latest.xlsx (and a dated archive) under data/output/workbook/.
    Returns the latest.xlsx path. Never raises on a missing date. The hourly
    job calls this with ``render_cards=True`` to bake in the card images."""
    latest, out = default_paths()
    out.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(data, render_cards=render_cards)
    wb.save(latest)
    date = primary_date or data.get("primary_date")
    if date:
        wb.save(out / f"{date}.xlsx")
    return latest


def build_default() -> Path | None:
    """Read data/output/latest.json and build the workbook. Returns the path,
    or None if there's no latest.json. Safe to call from the hourly job."""
    import json
    src = config.OUTPUT_DIR / "latest.json"
    if not src.exists():
        return None
    data = json.loads(src.read_text())
    return build_to_disk(data)
