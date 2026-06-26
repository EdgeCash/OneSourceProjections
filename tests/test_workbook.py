"""Tests for the daily editable workbook builder.

Covers: the in-cell formula math (mirrors project547.odds), the slate -> row
flattening + the EV-reconciliation that keeps the sheet in step with the
engine, the structural guarantees (model cells locked & numeric, odds cells
editable & formula-driven), and a real formula-evaluation pass when the
optional ``formulas`` engine is installed.
"""

import io
import math

import pytest
from openpyxl import load_workbook

from project547 import odds, workbook


# --------------------------------------------------------------------------
# Fixture slate: one fully-priced game (engine EVs present), one game with a
# moneyline only and no stored EV (raw-fallback path), and one priced prop.
# --------------------------------------------------------------------------
def _fixture():
    # Engine prices the away ML: 49.6% at +134 -> EV computed by the engine.
    ev_away = odds.expected_value(0.4961, 134)
    ev_home = odds.expected_value(0.5039, -136)
    return {
        "generated_at": "2026-06-25T22:00:00-04:00",
        "primary_date": "2026-06-25",
        "dates": ["2026-06-25"],
        "slates": {
            "2026-06-25": {
                "MLB": {
                    "games": [
                        {
                            "away_team": "Royals", "home_team": "Rays",
                            "game_time": "2026-06-25T16:10:00Z",
                            "away_win_prob": 0.5664, "home_win_prob": 0.4336,
                            "away_ml": 134, "home_ml": -136,
                            "away_ml_ev": ev_away, "home_ml_ev": ev_home,
                            "total_line": 8.5, "model_over_prob": 0.61,
                            "over_odds": 113, "under_odds": 105,
                            "over_ev": None, "under_ev": None,
                        },
                    ],
                    "props": [
                        {"player": "A. Judge", "team": "NYY", "opponent": "BOS",
                         "market": "batter_total_bases", "line": 1.5,
                         "odds": -120, "model_over_prob": 0.58},
                        {"player": "No Line", "market": "pitcher_strikeouts",
                         "line": None, "odds": None, "model_over_prob": None,
                         "projection": 5.5},
                    ],
                },
                "WNBA": {"games": [], "props": []},
            }
        },
        "performance": {
            "overall": {"graded_games": 208, "model_brier": 0.2482, "bets": 152,
                        "bet_win_rate": 0.4803, "units": 6.95, "roi_pct": 4.57,
                        "avg_clv_pct": 4.62, "clv_beat_rate": 0.5303},
            "by_sport": {"MLB": {"graded_games": 185, "model_brier": 0.2502,
                                 "bets": 130, "bet_win_rate": 0.47, "units": 4.0,
                                 "roi_pct": 3.1, "avg_clv_pct": 3.9,
                                 "clv_beat_rate": 0.52}},
        },
    }


# --------------------------------------------------------------------------
# Pure formula builders
# --------------------------------------------------------------------------
def test_formula_builders_noop_on_blank_odds():
    # every output formula must guard the empty odds cell
    assert workbook.implied_formula("$H$5").startswith('=IF($H$5="",""')
    assert workbook.ev_formula("$F$5", "$H$5").startswith('=IF($H$5="",""')
    assert workbook.kelly_pct_formula("$F$5", "$H$5").startswith('=IF($H$5="",""')
    assert workbook.stake_formula("$K$5", "$H$5").startswith('=IF($H$5="",""')


def test_verdict_thresholds_present():
    f = workbook.verdict_formula("$J$5")
    assert "Pass" in f and "Sharp band" in f and "Verify" in f
    # references the configurable min-edge setting, not a hard-coded number
    assert workbook.CELL_MIN_EDGE in f


def test_fair_american_matches_odds_module():
    assert workbook._fair_american(0.5) == odds.decimal_to_american(2.0)
    assert workbook._fair_american(0.0) is None  # degenerate
    assert workbook._fair_american(float("nan")) is None


# --------------------------------------------------------------------------
# Row flattening + EV reconciliation
# --------------------------------------------------------------------------
def test_game_rows_reconcile_to_engine_ev():
    g = _fixture()["slates"]["2026-06-25"]["MLB"]["games"][0]
    rows = workbook.game_rows("MLB", "2026-06-25", g)
    away = next(r for r in rows if r["selection"] == "Royals")
    # the probability shown must reproduce the engine's stored EV at the price
    assert away["engine_priced"] is True
    assert math.isclose(odds.expected_value(away["prob"], away["odds"]),
                        g["away_ml_ev"], abs_tol=1e-9)


def test_game_rows_raw_fallback_when_no_stored_ev():
    g = _fixture()["slates"]["2026-06-25"]["MLB"]["games"][0]
    rows = workbook.game_rows("MLB", "2026-06-25", g)
    over = next(r for r in rows if r["selection"].startswith("Over"))
    # over_ev was None -> fall back to the raw model prob, not engine-priced
    assert over["engine_priced"] is False
    assert math.isclose(over["prob"], 0.61)


def test_prop_rows_filter_unpriced():
    props = _fixture()["slates"]["2026-06-25"]["MLB"]["props"]
    rows = workbook.prop_rows("MLB", "2026-06-25", props)
    assert len(rows) == 1                      # the line-less prop is dropped
    assert rows[0]["player"] == "A. Judge"
    assert rows[0]["selection"] == "Over 1.5"


# --------------------------------------------------------------------------
# Structural guarantees on the built workbook
# --------------------------------------------------------------------------
def test_build_workbook_sheets():
    wb = workbook.build_workbook(_fixture())
    assert wb.sheetnames[:3] == ["Read Me", "Settings", "Top Plays"]
    assert "MLB Games" in wb.sheetnames
    assert "MLB Props" in wb.sheetnames
    assert "Track Record" in wb.sheetnames
    # empty WNBA slate produces no sheet
    assert "WNBA Games" not in wb.sheetnames


def test_odds_cell_editable_model_cell_locked():
    wb = workbook.build_workbook(_fixture())
    ws = wb["MLB Games"]
    assert ws.protection.sheet is True
    # H = your-odds (editable), F = model prob (locked & numeric)
    assert ws["H5"].protection.locked is False
    assert isinstance(ws["F5"].value, (int, float))
    assert ws["F5"].protection.locked is not False
    # the computed columns are genuine formulas, not baked numbers
    assert str(ws["J5"].value).startswith("=")   # EV
    assert str(ws["M5"].value).startswith("=")   # verdict


def test_top_plays_excludes_raw_fallback_rows():
    wb = workbook.build_workbook(_fixture())
    tp = wb["Top Plays"]
    sels = [tp.cell(row=r, column=5).value for r in range(5, tp.max_row + 1)]
    # the raw-fallback Over (fat but unpriced-by-engine) must not appear
    assert not any(s and str(s).startswith("Over") for s in sels)
    # the engine-priced away ML edge should
    assert "Royals" in sels


def test_build_bytes_roundtrips():
    data = workbook.build_workbook_bytes(_fixture())
    assert isinstance(data, bytes) and len(data) > 2000
    wb = load_workbook(io.BytesIO(data))
    assert "Top Plays" in wb.sheetnames


def test_track_record_has_overall_and_sport():
    wb = workbook.build_workbook(_fixture())
    ws = wb["Track Record"]
    labels = [ws.cell(row=r, column=1).value for r in range(5, 9)]
    assert "Overall" in labels and "MLB" in labels


# --------------------------------------------------------------------------
# Real formula evaluation (optional engine)
# --------------------------------------------------------------------------
def test_formulas_evaluate_correctly(tmp_path):
    formulas = pytest.importorskip("formulas")
    import warnings
    warnings.filterwarnings("ignore")
    p = tmp_path / "wb.xlsx"
    workbook.build_workbook(_fixture()).save(p)
    sol = formulas.ExcelModel().loads(str(p)).finish().calculate()

    def cell(addr):
        for k, v in sol.items():
            if k.upper().endswith(f"MLB GAMES'!{addr}"):
                try:
                    return float(v.value[0, 0])
                except Exception:
                    return v.value
        raise KeyError(addr)

    p_model, o = cell("F5"), cell("H5")
    assert math.isclose(cell("I5"), odds.implied_prob(o), abs_tol=1e-9)
    assert math.isclose(cell("J5"), odds.expected_value(p_model, o), abs_tol=1e-9)
    assert math.isclose(cell("K5"), odds.kelly_stake(p_model, o, 0.25), abs_tol=1e-9)
    assert math.isclose(cell("L5"), round(min(cell("K5"), 0.05) * 1000, 2), abs_tol=1e-2)
