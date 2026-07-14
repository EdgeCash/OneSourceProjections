"""Structured runner + daily Excel (.xlsx) workbook writer.

The daily deliverable is a **ready-to-play** workbook:

* one tab per DFS pick'em operator (PrizePicks, Underdog, Betr, Sleeper,
  Dabble) — the highest-probability over/under plays across every in-season
  sport, each with our model edge *and* BettingPros' second opinion;
* a Game Plays tab — moneyline / total / spread edges from the mature
  ``project547`` engine's committed ``data/output/latest.json``;
* Summary and Run_Log.

Excel, not Google Sheets, on purpose: openpyxl is already a repo dependency and
this needs no Google Cloud project, service account, or secrets. Run inside the
hourly job (keys present) so the BP/FP pulls it needs are cache-first and
bounded to once a day.

    python -m src.export --daily --source shared \
        --out data/output/latest.xlsx --json data/output/latest.json
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from . import api_client, db_manager, engine, grade

REPO_ROOT = Path(__file__).resolve().parents[2]


# --------------------------------------------------------------------------- #
# Shared runner
# --------------------------------------------------------------------------- #
def anchor_date(cli_date: str | None) -> str | None:
    """Slate date used to align the shared cache key with the main engine's
    pull: explicit --date, then the committed slate's ``primary_date``, then
    today ET."""
    if cli_date:
        return cli_date
    latest = REPO_ROOT / "data" / "output" / "latest.json"
    try:
        return json.loads(latest.read_text())["primary_date"]
    except Exception:
        pass
    try:
        from zoneinfo import ZoneInfo

        return datetime.now(ZoneInfo("America/New_York")).date().isoformat()
    except Exception:
        return None


def in_season_sports(date: str | None) -> list[str]:
    """In-season sports for the slate (falls back to MLB+WNBA offline)."""
    try:
        from project547.sports import active_sports

        return active_sports(date) or ["MLB", "WNBA"]
    except Exception:
        return ["MLB", "WNBA"]


def run_one(
    conn,
    sport: str,
    platform: str = "PrizePicks",
    budget: int = 50000,
    source: str = "shared",
    date: str | None = None,
    mode: str = "both",
) -> dict:
    """Run the engines for one sport / one operator; return a result dict."""
    sport = sport.upper()
    res: dict = {
        "sport": sport, "platform": platform, "budget": budget,
        "source": source, "dfs": [], "pickem": [],
        "proj_refreshed": 0, "lines_refreshed": 0, "lines_source": None,
    }
    if mode in ("dfs", "both"):
        res["proj_refreshed"] = api_client.refresh_projections(
            conn, sport, date=date, source=source)
        res["dfs"] = engine.optimize_salary_cap_dfs(sport, budget, conn=conn)
    if mode in ("pickem", "both"):
        res["lines_refreshed"] = api_client.refresh_market_lines(
            conn, sport, platform, source=source, date=date)
        res["pickem"] = engine.generate_optimal_pickem_slips(
            sport, platform, conn=conn)
        lines = db_manager.get_market_lines(conn, sport, platform)
        res["lines_source"] = (
            "bettingpros" if any(l.get("proj_mean") is not None for l in lines)
            else "derived" if lines else "none")
    return res


def collect_operators(conn, date, sports, source, per_op: int = 25) -> dict:
    """{operator: [ranked plays across all sports]} for every DFS operator."""
    for sport in sports:
        api_client.refresh_projections(conn, sport, date=date, source=source)

    operators: dict = {}
    for op in api_client.DFS_OPERATORS:            # PrizePicks, Underdog, ...
        plays: list = []
        for sport in sports:
            api_client.refresh_market_lines(conn, sport, op, source=source, date=date)
            plays += engine.generate_optimal_pickem_slips(
                sport, op, conn=conn, limit=per_op * 2)
        plays.sort(key=lambda p: p["confidence"], reverse=True)
        operators[op] = plays[:per_op]
    return operators


def game_plays(min_edge: float | None = None) -> list[dict]:
    """Moneyline / total / spread edges from the main engine's latest.json."""
    if min_edge is None:
        try:
            from project547 import config
            min_edge = config.MIN_EDGE
        except Exception:
            min_edge = 0.02
    try:
        data = json.loads((REPO_ROOT / "data" / "output" / "latest.json").read_text())
    except Exception:
        return []

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    out: list[dict] = []
    for _dt, sports in (data.get("slates") or {}).items():
        for sport, blob in (sports or {}).items():
            for g in (blob.get("games") if isinstance(blob, dict) else None) or []:
                matchup = f"{g.get('away_team')} @ {g.get('home_team')}"
                gt = g.get("game_time")
                mop = _f(g.get("model_over_prob"))
                mhc = _f(g.get("model_home_cover"))
                shl = _f(g.get("spread_home_line"))
                cands = [
                    ("Moneyline", g.get("home_team"), "", g.get("home_ml_ev"),
                     g.get("home_ml"), _f(g.get("home_win_prob"))),
                    ("Moneyline", g.get("away_team"), "", g.get("away_ml_ev"),
                     g.get("away_ml"), _f(g.get("away_win_prob"))),
                    ("Total", f"Over {g.get('total_line')}", "", g.get("over_ev"),
                     g.get("over_odds"), mop),
                    ("Total", f"Under {g.get('total_line')}", "", g.get("under_ev"),
                     g.get("under_odds"), None if mop is None else round(1 - mop, 4)),
                    ("Spread", f"{g.get('home_team')} {shl:+g}" if shl is not None
                     else g.get("home_team"), "", g.get("spread_home_ev"),
                     g.get("spread_home_odds"), mhc),
                    ("Spread", f"{g.get('away_team')} {-shl:+g}" if shl is not None
                     else g.get("away_team"), "", g.get("spread_away_ev"),
                     g.get("spread_away_odds"),
                     None if mhc is None else round(1 - mhc, 4)),
                ]
                for market, sel, _line, ev, odds, prob in cands:
                    ev = _f(ev)
                    if ev is None or ev < min_edge:
                        continue
                    out.append({
                        "sport": sport, "game": matchup, "market": market,
                        "selection": sel, "odds": _f(odds), "ev": round(ev, 4),
                        "model_prob": prob, "game_time": gt,
                    })
    out.sort(key=lambda p: p["ev"], reverse=True)
    return out


# --------------------------------------------------------------------------- #
# Excel workbook
# --------------------------------------------------------------------------- #
_HEADER_FILL = "1F2937"
_HEADER_FONT = "FFFFFF"
_BAND_FILL = "F3F4F6"
_TITLE_FONT = "111827"


def _style_header(ws, row_idx: int, ncols: int):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color=_HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)


def _autosize(ws, widths: dict[int, int]):
    from openpyxl.utils import get_column_letter

    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_table(ws, start_row: int, headers: list, rows: list[list],
                 fmts: dict[int, str] | None = None) -> int:
    from openpyxl.styles import PatternFill

    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header(ws, start_row, len(headers))
    band = PatternFill("solid", fgColor=_BAND_FILL)
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            if fmts and j in fmts:
                cell.number_format = fmts[j]
            if i % 2 == 1:
                cell.fill = band
        r += 1
    return r


def _title(ws, text: str, sub: str | None = None) -> int:
    from openpyxl.styles import Font

    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=15, color=_TITLE_FONT)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    return 4 if sub else 3


PCT, EDGE, ODDS, NUM, MONEY = "0.0%", "+0.0%;-0.0%", "+0;-0", "0.0", "#,##0"


def _odds_str(o, u) -> str:
    def a(v):
        return "" if v in (None, 0) else (f"+{int(v)}" if v > 0 else f"{int(v)}")
    return f"O {a(o)} / U {a(u)}".strip()


def write_workbook(payload: dict, logs: list[dict], out_path: str | Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()

    # --- Summary -----------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    row = _title(ws, "Daily Plays — ready to play",
                 f"generated {payload['generated_at']}  |  slate {payload.get('date')}"
                 f"  |  source: {payload['source']}")
    headers = ["Tab", "Plays", "Top play", "Top conf / EV"]
    rows = []
    for op, plays in payload["operators"].items():
        top = plays[0] if plays else None
        rows.append([op, len(plays),
                     f"{top['player_name']} {top['stat_type']} {top['side']} "
                     f"{top['line_value']}" if top else "-",
                     top["confidence"] if top else None])
    gp = payload["game_plays"]
    rows.append(["Game Plays", len(gp),
                 f"{gp[0]['selection']} ({gp[0]['market']})" if gp else "-",
                 f"{gp[0]['ev'] * 100:.1f}% EV" if gp else None])
    end = _write_table(ws, row, headers, rows)
    ws.cell(row=end + 1, column=1,
            value="Ranked by Confidence: our ensemble (model + BettingPros + "
                  "recent form) anchored to the sharp market, plus edge-vs-market, "
                  "signal agreement, and soft-line gap. Graded on closing-line "
                  "value. Personal research — not financial advice.")
    _autosize(ws, {1: 14, 2: 8, 3: 40, 4: 14})

    # --- One tab per DFS operator -----------------------------------------
    op_headers = ["#", "Sport", "Player", "Stat", "Line", "Side", "Conf",
                  "Win%", "Model%", "Edge vs Mkt", "Line edge", "Agree",
                  "BP EV", "BP★", "L10% O", "Odds (O/U)"]
    SIGNED = "+0.0;-0.0;0"
    for op, plays in payload["operators"].items():
        ws = wb.create_sheet(op[:31])
        row = _title(ws, f"{op} — pick'em board",
                     "ranked by our Confidence (ensemble edge vs the sharp "
                     "market + signal agreement + soft-line gap). Play the top.")
        rows = []
        for i, p in enumerate(plays, 1):
            rows.append([
                i, p["sport"], p["player_name"], p["stat_type"], p["line_value"],
                p["side"], p.get("confidence"), p["win_rate"], p.get("model_win"),
                p.get("edge_vs_market"), p.get("line_edge"),
                f"{p.get('agreement', 0)}/{p.get('n_signals', 0)}",
                p.get("bp_ev"), p.get("bet_rating"), p.get("form_l10"),
                _odds_str(p.get("over_odds"), p.get("under_odds")),
            ])
        if not rows:
            ws.cell(row=row, column=1, value="No plays for this operator today.")
        else:
            _write_table(ws, row, op_headers, rows,
                         fmts={5: NUM, 7: "0", 8: PCT, 9: PCT, 10: EDGE,
                               11: SIGNED, 13: EDGE, 14: "0.0", 15: PCT})
        _autosize(ws, {1: 4, 2: 6, 3: 22, 4: 13, 5: 6, 6: 6, 7: 6, 8: 8, 9: 8,
                       10: 11, 11: 9, 12: 7, 13: 8, 14: 6, 15: 8, 16: 15})

    # --- Game Plays --------------------------------------------------------
    ws = wb.create_sheet("Game Plays")
    row = _title(ws, "Game Plays — moneyline / total / spread",
                 "from the project547 engine (edges ≥ 2% EV), highest EV first")
    rows = [[p["sport"], p["game"], p["market"], p["selection"], p["odds"],
             p["model_prob"], p["ev"], p["game_time"]] for p in gp]
    if not rows:
        ws.cell(row=row, column=1, value="No qualifying game plays on this slate.")
    else:
        _write_table(ws, row, ["Sport", "Game", "Market", "Selection", "Odds",
                               "Model%", "EV", "Start (UTC)"],
                     rows, fmts={5: ODDS, 6: PCT, 7: EDGE})
    _autosize(ws, {1: 6, 2: 34, 3: 10, 4: 22, 5: 7, 6: 8, 7: 8, 8: 20})

    # --- Track Record ------------------------------------------------------
    tr = payload.get("track_record")
    if tr is not None:
        ws = wb.create_sheet("Track Record")
        row = _title(ws, "Track Record",
                     "graded on actual results — higher Confidence should hit "
                     "more, and we should beat BP when we disagree")
        if not tr.get("graded"):
            ws.cell(row=row, column=1,
                    value=f"No graded picks yet — {tr.get('pending', 0)} pending "
                          "until games are final.")
        else:
            ov = tr["overall"]

            def _pct(v):
                return "-" if v is None else f"{v:.1%}"

            ws.cell(row=row, column=1,
                    value=f"Overall: {ov['w']}-{ov['l']}-{ov['p']}  |  Hit "
                          f"{_pct(ov['hit'])}  |  vs 54.3% break-even: "
                          f"{'-' if ov['hit'] is None else format(ov['hit'] - 0.543, '+.1%')}"
                          f"  ({ov['n']} graded)")
            row += 2
            row = _write_table(ws, row, ["Confidence", "N", "W", "L", "Hit%"],
                               [[b["bucket"], b["n"], b["w"], b["l"], b["hit"]]
                                for b in tr["by_confidence"]], fmts={5: PCT})
            row += 1
            row = _write_table(ws, row, ["Operator", "N", "W", "L", "Hit%"],
                               [[o["operator"], o["n"], o["w"], o["l"], o["hit"]]
                                for o in tr["by_operator"]], fmts={5: PCT})
            row += 1
            vb = tr["vs_bp"]
            _write_table(ws, row, ["vs BettingPros", "N", "W", "L", "Hit%"],
                         [["Agreed with BP", vb["agree"]["n"], vb["agree"]["w"],
                           vb["agree"]["l"], vb["agree"]["hit"]],
                          ["Disagreed with BP", vb["disagree"]["n"], vb["disagree"]["w"],
                           vb["disagree"]["l"], vb["disagree"]["hit"]]], fmts={5: PCT})
        _autosize(ws, {1: 20, 2: 6, 3: 6, 4: 6, 5: 8})

    # --- Run_Log -----------------------------------------------------------
    ws = wb.create_sheet("Run_Log")
    row = _title(ws, "Run Log", "cache/source ledger — real BP/FP spend is "
                 "cache-first and bounded to this once-a-day build")
    rows = [[lg["timestamp"], lg["endpoint"], lg["request_count"]] for lg in logs]
    _write_table(ws, row, ["Timestamp", "Endpoint", "Requests"], rows)
    _autosize(ws, {1: 22, 2: 44, 3: 10})

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# --------------------------------------------------------------------------- #
# Picks history — the track record (foundation for closing-line-value grading)
# --------------------------------------------------------------------------- #
def append_picks_history(payload: dict, path: str | Path) -> int:
    """Append the day's pick'em plays to a committed JSONL, deduped by slate
    date, so the CLV track record accrues from day one. The grader (next build)
    reads this and scores each pick's line against the closing line.
    """
    path = Path(path)
    date = payload.get("date") or ""
    stamp = payload["generated_at"]
    new_rows = []
    for op, plays in payload["operators"].items():
        for p in plays:
            new_rows.append({
                "date": date, "logged_at": stamp, "operator": op,
                "sport": p["sport"], "player": p["player_name"],
                "stat": p["stat_type"], "line": p["line_value"], "side": p["side"],
                "win_rate": p["win_rate"], "model_win": p.get("model_win"),
                "confidence": p.get("confidence"),
                "edge_vs_market": p.get("edge_vs_market"),
                "line_edge": p.get("line_edge"), "consensus_line": p.get("consensus_line"),
                "over_odds": p.get("over_odds"), "under_odds": p.get("under_odds"),
                "bp_ev": p.get("bp_ev"), "bp_recommended": p.get("bp_recommended"),
                # graded later:
                "closing_line": None, "closing_over_odds": None,
                "closing_under_odds": None, "clv": None, "result": None,
            })
    existing = []
    if path.exists():
        for ln in path.read_text().splitlines():
            if not ln.strip():
                continue
            try:
                row = json.loads(ln)
            except ValueError:
                continue
            if row.get("date") != date:      # replace this slate's block on re-run
                existing.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(r, default=str)
                              for r in existing + new_rows) + "\n")
    return len(new_rows)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Build the daily ready-to-play workbook")
    ap.add_argument("--daily", action="store_true",
                    help="auto in-season sports + all DFS operators + game plays")
    ap.add_argument("--sports", default=None, help="comma-separated; default: in-season")
    ap.add_argument("--source", choices=["shared", "sample"], default="shared")
    ap.add_argument("--date", default=None)
    ap.add_argument("--out", default="data/output/latest.xlsx")
    ap.add_argument("--json", default=None)
    args = ap.parse_args(argv)

    date = anchor_date(args.date)
    sports = ([s.strip().upper() for s in args.sports.split(",") if s.strip()]
              if args.sports else in_season_sports(date))

    conn = db_manager.connect()
    db_manager.init_db(conn)
    try:
        operators = collect_operators(conn, date, sports, args.source)
        gp = game_plays()
        logs = [dict(r) for r in conn.execute(
            "SELECT timestamp, endpoint, request_count FROM api_log ORDER BY id")]
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source": args.source, "date": date, "sports": sports,
            "budget_used": db_manager.api_usage_today(conn),
            "daily_budget": api_client.DAILY_BUDGET,
            "operators": operators, "game_plays": gp,
        }
        hist_path = Path(args.out).parent / "picks_history.jsonl"
        hist = append_picks_history(payload, hist_path)
        graded = grade.grade_history(hist_path)       # grade any now-final games
        payload["track_record"] = grade.summarize(hist_path)
        out = write_workbook(payload, logs, args.out)
        msg = f"wrote {out} ({sum(len(v) for v in operators.values())} pick'em plays, " \
              f"{len(gp)} game plays across {', '.join(sports)}; logged {hist} picks, " \
              f"graded {graded['graded']} ({graded['wins']}-{graded['losses']}-{graded['pushes']}))"
        if args.json:
            jp = Path(args.json)
            jp.parent.mkdir(parents=True, exist_ok=True)
            jp.write_text(json.dumps(payload, indent=1, default=str))
            msg += f" and {jp}"
        print(msg)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
